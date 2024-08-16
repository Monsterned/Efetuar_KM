import pandas as pd
from openpyxl import load_workbook
import os
import numpy as np
import pyautogui
import ctypes
from datetime import datetime
import pyautogui
import shutil
from datetime import datetime, timedelta

caminho = os.getcwd() 
caminho_sistema = caminho.replace("C", "T", 1)

def click_image(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        image_path2 = os.path.join(current_dir, caminho_imagem, "query_timeout_expered.png") 
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Imagem de query_timeout_expered foi encontrada na tela.")
                click_image('ok.png')
        except Exception as e:
            print("Aguardando...")
        pyautogui.sleep(1)

def campo_data_baixa(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def tela_baixa(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            clique_baixa('sim_cancelamento.png')
            clique_baixa('yes.png')           
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def clique_baixa(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            click('baixar_manifesto.png')
            print("Clicou novamente na baixa.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def sim_reculculo(image_path,image_path2, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                center_x = position2.left + position2.width // 2
                center_y = position2.top + position2.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def click(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def click_encerrar(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.click(center_x, center_y)
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            clicar_encerrar('ok.png')
            clicar_encerrar('yes.png')
            pyautogui.sleep(1)
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def clicar_encerrar(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            center_x = position.left + position.width // 2
            center_y = position.top + position.height // 2
            pyautogui.click(center_x, center_y)
            click('encerrar_sefaz.png')
            print("Imagem foi encontrada na tela.")
    except Exception as e:
        print("Imagem não encontrada na tela. Aguardando...")
    pyautogui.sleep(1)

def novo_lançamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        image_path2 = os.path.join(current_dir, caminho_imagem, "query_timeout_expered.png") 
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Imagem de query_timeout_expered foi encontrada na tela.")
                click_image('ok.png')
        except Exception as e:
            print("Aguardando...")
        pyautogui.sleep(1)

def click_info_manifesto(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                center_x = position.left + position.width // 2
                center_y = position.top + position.height // 2
                pyautogui.moveTo(center_x, center_y)  # Movendo o cursor para a posição da imagem               
                pyautogui.moveRel(30, 0)  # Movendo o cursor para cima
                pyautogui.click()  # Clicando no local da imagem
                print("Imagem foi encontrada na tela.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        image_path2 = os.path.join(current_dir, caminho_imagem, "query_timeout_expered.png") 
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Imagem de query_timeout_expered foi encontrada na tela.")
                click_image('ok.png')
        except Exception as e:
            print("Aguardando...")
        pyautogui.sleep(1)

def confirmação(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    not_found_count = 0
    max_attempts = 3
    while not_found_count < max_attempts:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Imagem foi encontrada na tela.")
                not_found_count = 0  # Reseta o contador se a imagem for encontrada
            else:
                not_found_count += 1
                print(f"Imagem não encontrada na tela. Tentativa {not_found_count} de {max_attempts}.")
        except Exception as e:
            print(f"Erro ao tentar localizar a imagem: {e}")
            not_found_count += 1
        if not_found_count < max_attempts:
            pyautogui.sleep(1)
        else:
            print("Imagem não encontrada após 3 tentativas. Saindo da função.")
            break

def verificar_campo(image_path,image_path2,image_path3, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path2 = os.path.join(current_dir, caminho_imagem, image_path2) 
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                print("Campo encontrado.")
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        pyautogui.sleep(1)

def verificar_campo_cancelamento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Campo de cancelamento encontrado.")
                break
        except Exception as e:
            print("Campo de cancelamento nao encontrado. Aguardando...")
        pyautogui.sleep(1)

def check_caps_lock():
    return ctypes.windll.user32.GetKeyState(0x14) & 0xffff != 0

def aviso_vencimento(image_path, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = os.path.join(current_dir, 'IMAGENS')  # Correção do caminho
    image_path = os.path.join(caminho_imagem, image_path)
    aviso_ativo = False
    try:
        position = pyautogui.locateOnScreen(image_path, confidence=confidence)
        if position:
            click_image('ok_marcado.png')
            print("Imagem foi encontrada na tela.")
            aviso_ativo = True
    except Exception as e:
        print(f"Erro ao tentar localizar a imagem: {e}")

    return aviso_ativo

def status_manifesto(image_path,image_path3,image_path4, confidence=0.9):
    current_dir = os.path.dirname(__file__)  # Diretório atual do script
    caminho_imagem = caminho + r'\IMAGENS'
    image_path = os.path.join(current_dir, caminho_imagem, image_path) 
    image_path3 = os.path.join(current_dir, caminho_imagem, image_path3)
    image_path4 = os.path.join(current_dir, caminho_imagem, image_path4)
    nao_encontrado = False
    aviso_ativo = False
    while True:
        try:
            position = pyautogui.locateOnScreen(image_path, confidence=confidence)
            if position:
                print("Campo encontrado cadastrado.")
                aviso_ativo = aviso_vencimento('curso_vencido.png')
                return aviso_ativo,nao_encontrado
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            position3 = pyautogui.locateOnScreen(image_path3, confidence=confidence)
            if position3:
                print("Campo encontrado emitido.")
                aviso_ativo = aviso_vencimento('curso_vencido.png')
                return aviso_ativo,nao_encontrado
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        image_name2 = 'status_baixado.png'
        image_path2 = os.path.join(current_dir, caminho_imagem, image_name2)
        try:
            position4 = pyautogui.locateOnScreen(image_path4, confidence=confidence)
            if position4:
                print("Manifesto nao localizado.")
                nao_encontrado = True
                caminho_do_arquivo = 'BASE_KM.xlsx'
                nome_da_aba = 'Planilha1'
                wb = load_workbook(caminho_do_arquivo)
                ws = wb[nome_da_aba]
                coluna_ost = 'E'  
                if linha > ws.max_row:
                    ws[coluna_ost + str(linha_especifica)] = 'NAO LOCALIZADO'
                else:
                    ws[coluna_ost + str(linha_especifica)] = 'NAO LOCALIZADO'
                wb.save(caminho_do_arquivo)
                wb.close()
                return aviso_ativo,nao_encontrado
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")
        try:
            print("Tentando achar a segunda imagem.")
            position2 = pyautogui.locateOnScreen(image_path2, confidence=confidence)
            if position2:
                print("Campo de baixado encontrado.")
                aviso_ativo = aviso_vencimento('curso_vencido.png')
                click_image('botao_faturamento.png')
                click_image('botao_faturamento_movimentacao.png')
                click_image('botao_faturamento_movimentacao_cancelamento.png')
                pyautogui.sleep(2)
                for i in range(2):
                    pyautogui.press("enter")
                pyautogui.sleep(1)
                click_image('janela.png')
                click_image('tela_cancelamento.png')
                click_image('confirmacao_tela_cancelamento.png')
                click_image('incluir.png')
                pyautogui.sleep(2)
                verificar_campo_cancelamento('status_tela_cancelamento.png')
                click_image('CTRC.png')
                for i in range(10): 
                    pyautogui.press("up")
                #click_image('cancelar_baixa_manifesto.png')
                pyautogui.press("down")  
                pyautogui.press("enter")   
                pyautogui.sleep(2)
                click_info_manifesto('filial_cancelamento.png')
                pyautogui.press("backspace")
                pyautogui.write(str(filial))
                pyautogui.sleep(1)
                pyautogui.press("tab")
                click_info_manifesto('serie_cancelamento.png')
                pyautogui.press("backspace")
                pyautogui.write(str(serie))
                pyautogui.sleep(1)
                click_info_manifesto('manifesto_cancelamento.png')
                for i in range(10): 
                    pyautogui.press("backspace")
                pyautogui.write(str(manifesto))
                pyautogui.sleep(1)
                for i in range(2):
                    pyautogui.press("tab")
                pyautogui.press("1")
                pyautogui.sleep(1)
                pyautogui.press("tab")
                pyautogui.write("correcao de km")
                pyautogui.sleep(1)
                click_image('salvar_cancelamento.png')
                click_image('sim_cancelamento.png')
                click_image('ok_cancelamento.png')
                click_image('voltar.png')
                print('cancelado')
                pyautogui.sleep(2)
                click_image('incluir.png')
                pyautogui.sleep(2)
                novo_lançamento('campo_numero_manifesto.png')
                pyautogui.sleep(1)
                click_info_manifesto('filial.png')
                pyautogui.press("backspace")
                pyautogui.write(str(filial))
                pyautogui.sleep(1)
                pyautogui.press("tab")
                click_info_manifesto('serie.png')
                pyautogui.press("backspace")
                pyautogui.write(str(serie))
                pyautogui.sleep(1)
                pyautogui.press("tab")
                click_info_manifesto('campo_manifesto.png')
                for i in range(10): 
                    pyautogui.press("backspace")
                for i in range(10): 
                    pyautogui.press("del")
                pyautogui.sleep(0.5)
                print(manifesto)
                pyautogui.write(str(manifesto))
                pyautogui.sleep(1)
                pyautogui.press("tab")
                status_manifesto('status_cadastrado.png','status_emitido.png')
                return aviso_ativo,nao_encontrado
                break
        except Exception as e:
            print("Imagem não encontrada na tela. Aguardando...")

        pyautogui.sleep(1)
    
def check_caps_lock():
    return ctypes.windll.user32.GetKeyState(0x14) & 0xffff != 0

def alt_press(key):
    pyautogui.keyDown('alt')
    pyautogui.press(key)
    pyautogui.keyUp('alt')

click_image('botao_faturamento.png')
click_image('botao_faturamento_movimentacao.png')
click_image('botao_faturamento_movimentacao_manifesto.png')
click_image('botao_faturamento_movimentacao_manifesto_emissao.png')

Planilha_km = pd.read_excel("BASE_KM.xlsx")
#print(Planilha_km)

numero_linhas = len(Planilha_km)
linha_especifica = 1
for i, linha in enumerate(Planilha_km.index):
    filial = Planilha_km.loc[linha, "FILIAL"]
    filial = int(filial)
    serie = Planilha_km.loc[linha, "SERIE"]
    serie = int(serie)
    manifesto = Planilha_km.loc[linha, "MANIFESTO"]
    km = Planilha_km.loc[linha, "KM"]
    km = int(km)
    agora = datetime.now()
    um_minuto_atras = agora - timedelta(minutes=1)
    agora_formatado = agora.strftime("%d/%m/%Y %H:%M")
    um_minuto_atras_formatado = um_minuto_atras.strftime("%d/%m/%Y%H:%M")
    falta = numero_linhas - i
    linha_especifica += 1
    print(f'filial:{filial} serie:{serie} manifesto:{manifesto} km:{km} falta:{falta}')
    click_image('incluir.png')
    pyautogui.sleep(2)
    novo_lançamento('campo_numero_manifesto.png')
    click_info_manifesto('filial.png')
    pyautogui.press("backspace")
    pyautogui.write(str(filial))
    pyautogui.sleep(1)
    pyautogui.press("tab")
    click_info_manifesto('serie.png')
    pyautogui.press("backspace")
    pyautogui.write(str(serie))
    pyautogui.sleep(1)
    pyautogui.press("tab")
    click_info_manifesto('campo_manifesto.png')
    print('ta aqui')
    for i in range(10): 
        pyautogui.press("backspace")
    for i in range(10): 
        pyautogui.press("del")
    pyautogui.sleep(0.5)
    pyautogui.write(str(manifesto))
    pyautogui.sleep(1)
    pyautogui.press("tab")
    for i in range(2):
        pyautogui.press("enter")
    aviso_ativo,nao_encontrado = status_manifesto('status_cadastrado.png','status_emitido.png','campo_numero_manifesto.png')
    pyautogui.sleep(1)
    if nao_encontrado:
        continue
    for i in range(2):
        pyautogui.press("enter")
    click_info_manifesto('campo_km.png')
    for i in range(10): 
        pyautogui.press("backspace")
    for i in range(10): 
        pyautogui.press("del")
    pyautogui.write(str(km))
    pyautogui.sleep(1)
    click_image('salvar_km.png')
    click_image('ok.png')
    pyautogui.sleep(2)
    click_image('botao_recalculo.png')
    sim_reculculo('sim_recalculo.png','ok.png')
    pyautogui.sleep(3)
    click_image('baixar_manifesto.png')
    tela_baixa('tela_baixar_data.png')
    campo_data_baixa('campo_data_baixa.png','campo_data_baixa2.png')
    pyautogui.sleep(1)
    for i in range(16):
        pyautogui.press("backspace")
    pyautogui.write(str(um_minuto_atras_formatado))
    pyautogui.sleep(1)
    click_image('ok2.png')
    pyautogui.sleep(1)
    click_image('ok.png')
    pyautogui.sleep(1)
    click_image('nao.png')
    pyautogui.sleep(2)

    if aviso_ativo:
        click_image('ok.png')
        pyautogui.sleep(1)
    click_image('encerrar_sefaz.png')
    click_encerrar('campo_encerramento_numero_manifesto.png')
    pyautogui.press('enter')
    pyautogui.write(str(manifesto))
    click_image('botao_encerrar.png')
    pyautogui.sleep(4)
    verificar_campo('encerrado.png','manifesto_ja_baixado.png','encerrado_secretaria.png')
    click_image('fechar_tela_encerramento.png')
    pyautogui.sleep(2)

    caminho_do_arquivo = 'BASE_KM.xlsx'
    nome_da_aba = 'Planilha1'
    wb = load_workbook(caminho_do_arquivo)
    ws = wb[nome_da_aba]
    coluna_ost = 'E'  
    if linha > ws.max_row:
        ws[coluna_ost + str(linha_especifica)] = 'MANIFESTO BAIXADO'
    else:
        ws[coluna_ost + str(linha_especifica)] = 'MANIFESTO BAIXADO'
    wb.save(caminho_do_arquivo)
    wb.close()

pyautogui.sleep(2)
click_image('voltar.png')
